import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import Handle_file as hf
import openpyxl
import os
import json
import datetime
import webbrowser
import tempfile
import subprocess

class CustomApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Customizable Application")
        self.geometry("1000x700")
        self.configure(bg="#4786E6")
        self.checklist_count = 0
        self.tree = None
        self.excel_path = None
        self.headers = []
        self.data = []
        self.project = "DAS_VW_02"
        self.working_dir = None
        self.Test_level = "30_SW_Test"
    
        # PanedWindow cho phép resize các vùng
        self.paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, bg="#4786E6")
        self.paned.pack(fill="both", expand=True)

        # Sidebar
        self.sidebar = tk.Frame(self.paned, bg="#077786", width=200)
        self.sidebar.pack_propagate(False)
        tk.Label(self.sidebar, text="Công cụ", bg="#077786", fg="white", font=("Segoe UI", 12, "bold")).pack(pady=5)

        # Group 1: Quản lý Excel
        frame_excel = tk.LabelFrame(self.sidebar, text="Excel mannagment", bg="#077786", fg="white")
        frame_excel.pack(fill="x", padx=8, pady=4)
        ttk.Button(frame_excel, text="Open new Excel file", command=self.load_excel_table).pack(fill="x", pady=2)
        ttk.Button(frame_excel, text="Save Excel file", command=self.save_excel_table).pack(fill="x", pady=2)
        ttk.Button(frame_excel, text="Open Excel file", command=self.open_excel_file).pack(fill="x", pady=2)

        # Group 2: Chỉnh sửa bảng
        frame_edit = tk.LabelFrame(self.sidebar, text="Table setting", bg="#077786", fg="white")
        frame_edit.pack(fill="x", padx=8, pady=4)
        ttk.Button(frame_edit, text="Add column", command=self.add_column_to_table).pack(fill="x", pady=2)
        ttk.Button(frame_edit, text="Add row", command=self.add_row_to_table).pack(fill="x", pady=2)

        # Group 3: Quản lý workspace
        frame_ws = tk.LabelFrame(self.sidebar, text="Workspace", bg="#077786", fg="white")
        frame_ws.pack(fill="x", padx=8, pady=4)
        ttk.Button(frame_ws, text="Open new workspace", command=self.load_workspace).pack(fill="x", pady=2)
        ttk.Button(frame_ws, text="Save workspace", command=self.save_workspace).pack(fill="x", pady=2)
        ttk.Button(frame_ws, text="Save as workspace", command=lambda: self.save_workspace(save_as=True)).pack(fill="x", pady=2)
        

        # Group 4: Khác
        frame_other = tk.LabelFrame(self.sidebar, text="Checking file", bg="#077786", fg="white")
        frame_other.pack(fill="x", padx=8, pady=4)
        self.btn_open_script = ttk.Button(frame_other, text="Open Automation script", command=self.open_script)
        self.btn_open_script.pack(fill="x", pady=2)
        ttk.Button(frame_other, text="Check Consistency", command=self.compare_content_requirement).pack(fill="x", pady=2)

        # Group 5: Khác
        frame_refresh = tk.LabelFrame(self.sidebar, text="", bg="#077786", fg="white")
        frame_refresh.pack(side= "bottom", fill="x", padx=8, pady=4)
        ttk.Button(frame_refresh, text="Save all & refresh", command=self.refresh_all).pack(fill="x", pady=2)

        self.paned.add(self.sidebar, minsize=120)

        # Main area
        self.main_area = tk.Frame(self.paned, bg="white")
        self.paned.add(self.main_area, minsize=300)
        tk.Label(self.main_area, text="Main Area", bg="white", font=("Segoe UI", 14, "bold")).pack(pady=10)

        # Footer (cố định dưới cùng, không resize)
        self.footer = tk.Frame(self, bg="#e0e0e0", height=20)
        self.footer.pack(side="bottom", fill="x")
        tk.Label(self.footer, text="Develop by Nguyen Loc", bg="#e0e0e0").pack(pady=1)
        self.status_label = tk.Label(self.footer, text="", bg="#e0e0e0", fg="green", anchor="e", font=("Segoe UI", 10, "bold"))
        self.status_label.pack(side="right", padx=10)

        # Đăng ký style màu xanh cho Success.TButton
        style = ttk.Style()
        style.configure("Success.TButton", foreground="Green", background="#11e61b")
        style.map("Success.TButton",
          background=[('active', '#388e3c'), ('!active', '#2e7d32')])

    def add_button_to_main(self):
        ttk.Button(self.main_area, text="Nút mới").pack(pady=5)

    def handle_add_checklist(self):
        self.checklist_count += 1
        self.add_checklist_to_main(f"Checklist {self.checklist_count}")

    def add_checklist_to_main(self, name_of_option):
        frame = tk.Frame(self.main_area, bg="white")
        frame.pack(pady=5)
        tk.Label(frame, text="Checklist:", bg="white").pack(side="left")
        var = tk.BooleanVar()
        tk.Checkbutton(frame, text=name_of_option, variable=var, bg="white").pack(side="left")

    def load_excel_table(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        self.excel_path = file_path
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Đảm bảo header là chuỗi, không None, không rỗng, không trùng lặp
        self.headers = []
        for i, cell in enumerate(ws[1]):
            val = cell.value if cell.value not in [None, ""] else f"Cột {i+1}"
            while val in self.headers:
                val += "_1"
            self.headers.append(str(val))
        self.data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.data.append(list(row))
        wb.close()
        self.refresh_table()

    def on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return
        col_idx = int(column.replace("#", "")) - 1
        row_idx = self.tree.index(item)
        x, y, width, height = self.tree.bbox(item, column)
        value = self.tree.set(item, column)

        entry = tk.Entry(self.tree)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus()

        def save_edit(event):
            new_value = entry.get()
            self.tree.set(item, column, new_value)
            self.data[row_idx][col_idx] = new_value
            entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", lambda e: entry.destroy())

    def save_excel_table(self):
        if not self.excel_path or not self.data:
            messagebox.showwarning("Chưa có dữ liệu", "Hãy mở file Excel trước!")
            return
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        # Ghi header vào dòng 1
        for j, header in enumerate(self.headers, start=1):
            ws.cell(row=1, column=j, value=header)
        # Ghi dữ liệu vào các dòng tiếp theo
        for i, row in enumerate(self.data, start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
        wb.save(self.excel_path)
        wb.close()
        self.set_status("Đã lưu", "Dữ liệu đã được lưu vào file Excel!")

    def open_excel_file(self):
        if not self.excel_path:
            messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
            return
        hf.open_excel_file("", self.excel_path)

    def add_column_to_table(self):
        if not self.excel_path:
            messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
            return
        # Hỏi tên cột mới
        new_col_name = simpledialog.askstring("Đổi tên cột", "Nhập tên cột mới:", initialvalue=f"Cột mới {len(self.headers) + 1}")
        if not new_col_name:
            new_col_name = f"Cột mới {len(self.headers) + 1}"
        # Đảm bảo không trùng tên
        while new_col_name in self.headers:
            new_col_name += "_1"
        self.headers.append(new_col_name)
        for row in self.data:
            row.append("")
        self.refresh_table()

    def add_row_to_table(self):
        if not self.excel_path:
            messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
            return
        new_row = [""] * len(self.headers)
        self.data.append(new_row)
        self.refresh_table()

    def refresh_table(self):
        for widget in self.main_area.winfo_children():
            if isinstance(widget, ttk.Treeview):
                widget.destroy()
        # Chỉ lọc các cột có tên (không rỗng/None)
        valid_indices = []
        valid_headers = []
        for idx, h in enumerate(self.headers):
            if h not in [None, ""]:
                valid_indices.append(idx)
                valid_headers.append(h)
        self.headers = valid_headers
        filtered_data = []
        for row in self.data:
            filtered_row = [row[idx] if idx < len(row) and row[idx] is not None else "" for idx in valid_indices]
            filtered_data.append(filtered_row)
        self.data = filtered_data

        self.tree = ttk.Treeview(self.main_area, columns=self.headers, show="headings")
        # Định nghĩa tag màu
        self.tree.tag_configure('passed', background='#b6fcb6')      # Xanh lá nhạt
        self.tree.tag_configure('failed', background='#ffb3b3')      # Đỏ nhạt
        self.tree.tag_configure('not_tested', background='#fff7b2')  # Vàng nhạt
        self.tree.tag_configure('discarded', background='#fff7b2')   # Vàng nhạt

        for col in self.headers:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        # --- Thêm dấu tick khi hiển thị ---
        try:
            crid_idx = self.headers.index("CR Related")
            feature_idx = self.headers.index("Feature")
            id_idx = self.headers.index("Test cases ID")
        except Exception:
            crid_idx = feature_idx = id_idx = None

        for row in self.data:
            display_row = [cell if cell is not None else "" for cell in row]
            # Thêm dấu tick nếu file đã tồn tại
            if crid_idx is not None and feature_idx is not None and id_idx is not None and self.working_dir:
                crid_val = str(display_row[crid_idx]).strip()
                feature_val_raw = str(display_row[feature_idx]).strip()
                id_val = str(display_row[id_idx]).strip()
                feature_map = {
                    "Cust": "Customization", "cust": "Customization", "Customization": "Customization",
                    "Norm": "Normalization", "norm": "Normalization", "Normalization": "Normalization",
                    "Diag": "UDSDiagnostics", "diag": "UDSDiagnostics", "UDSDiagnostics": "UDSDiagnostics",
                    "DTC": "DTCandErrorHandling", "dtc": "DTCandErrorHandling",
                    "ProgramSequenceMonitoring": "ProgramSequenceMonitoring", "PSM": "ProgramSequenceMonitoring"
                }
                feature_val = feature_map.get(feature_val_raw, feature_val_raw)
                crid_folder = os.path.join(self.working_dir, crid_val)
                feature_folder = os.path.join(crid_folder, feature_val)
                save_path = os.path.join(feature_folder, f"{id_val}.can")
                if os.path.exists(save_path) and id_idx < len(display_row):
                    display_row[id_idx] = f"{display_row[id_idx]} ✅"
            # Xác định tag theo giá trị cột "Test Verdict"
            tag = ""
            try:
                verdict_idx = self.headers.index("Test Verdict")
                verdict = str(display_row[verdict_idx]).strip().lower()
                if verdict == "passed":
                    tag = "passed"
                elif verdict == "failed":
                    tag = "failed"
                elif verdict in ("not_tested", "discarded"):
                    tag = verdict
            except Exception:
                pass
            self.tree.insert("", "end", values=display_row, tags=(tag,))
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.bind("<Double-1>", self.on_double_click)

    def save_workspace(self, save_as=False):
        def convert_value(val):
            if isinstance(val, datetime.datetime):
                return val.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(val, datetime.date):
                return val.strftime("%Y-%m-%d")
            return val

        serializable_data = [
            [convert_value(cell) for cell in row]
            for row in self.data
        ]
        workspace = {
            "headers": self.headers,
            "data": serializable_data,
            "sidebar_width": self.sidebar.winfo_width(),
            "main_area_width": self.main_area.winfo_width(),
            "excel_path": self.excel_path,
            "working path": self.working_dir
        }

        # Nếu đã có workspace_path và không yêu cầu lưu mới, thì lưu đè
        if hasattr(self, "workspace_path") and self.workspace_path and not save_as:
            file_path = self.workspace_path
        else:
            file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
            if not file_path:
                self.set_status("Hủy lưu workspace.", success=False)
                return
            self.workspace_path = file_path

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(workspace, f)
        self.set_status("Đã lưu workspace thành công!", success=True)

    def load_workspace(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                workspace_data = json.load(f)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc workspace:\n{e}")
            return
        self.workspace_path = file_path  # Ghi nhớ đường dẫn workspace
        self.excel_path = workspace_data.get("excel_path")
        self.headers = workspace_data.get("headers", [])
        self.data = workspace_data.get("data", [])
        self.sidebar_width = workspace_data.get("sidebar_width", 200)
        self.main_area_width = workspace_data.get("main_area_width", 300)
        self.working_dir = workspace_data.get("working path", None)
        self.refresh_table()
        self.set_status("Workspace đã được mở!", success=True)
        # messagebox.showinfo("Đã mở", "Workspace đã được mở!")
    
    def open_script(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showwarning("Chọn dòng", "Hãy chọn một dòng trong bảng!")
            return
        values = self.tree.item(selected, "values")
        self.ensure_script_file(values, auto_open=True)

    def ensure_script_file(self, values, auto_open=False):
        """Đảm bảo file .can đã có ở thư mục working, nếu chưa thì tải/copy về. 
        Nếu auto_open=True thì mở file sau khi copy."""
        try:
            crid_idx = self.headers.index("CR Related")
            feature_idx = self.headers.index("Feature")
            id_idx = self.headers.index("Test cases ID")
        except ValueError:
            messagebox.showerror("Lỗi", "Không tìm thấy cột 'CR ID', 'Feature' hoặc 'Test cases ID'!")
            return None

        crid_val = str(values[crid_idx]).strip()
        feature_val_raw = str(values[feature_idx]).strip()
        id_val = str(values[id_idx]).strip()
        feature_map = {
            "Cust": "Customization", "cust": "Customization", "Customization": "Customization",
            "Norm": "Normalization", "norm": "Normalization", "Normalization": "Normalization",
            "Diag": "UDSDiagnostics", "diag": "UDSDiagnostics", "UDSDiagnostics": "UDSDiagnostics",
            "DTC": "DTCandErrorHandling", "dtc": "DTCandErrorHandling",
            "ProgramSequenceMonitoring": "ProgramSequenceMonitoring", "PSM": "ProgramSequenceMonitoring"
        }
        feature_val = feature_map.get(feature_val_raw, feature_val_raw)

        # Chọn/tham chiếu thư mục working
        if not self.working_dir:
            self.working_dir = filedialog.askdirectory(title="Chọn thư mục working để lưu script")
            if not self.working_dir:
                messagebox.showwarning("Chưa chọn thư mục", "Hãy chọn thư mục để lưu script!")
                return None
        crid_folder = os.path.join(self.working_dir, crid_val)
        feature_folder = os.path.join(crid_folder, feature_val)
        os.makedirs(feature_folder, exist_ok=True)
        save_path = os.path.join(feature_folder, f"{id_val}.can")

        # Nếu file đã tồn tại thì trả về luôn
        if os.path.exists(save_path):
            if auto_open:
                try:
                    os.startfile(save_path)
                    self.set_status(f"Đã mở file: {save_path}", success=True)
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể mở file {save_path}: {e}")
                    return None
            return save_path

        # Nếu chưa có file thì tải về và copy
        url = (
            f"http://mks1.dc.hella.com:7001/si/viewrevision?"
            f"projectName=e:/Projects/DAS_RADAR/30_PRJ/10_CUST/10_VAG/{self.project}/60_ST/{self.Test_level}/20_SWT_CC/10_Debugger_Test/20_Scripts/Test_Cases/"
            f"{feature_val}/project.pj&selection={id_val}.can&revision=:member"
        )
        webbrowser.open(url)
        self.set_status("Đang tải file, vui lòng tải file về thư mục Download...", success=True)
        download_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        src_file = os.path.join(download_dir, id_val + ".can")
        for _ in range(30):
            if os.path.exists(src_file):
                break
            self.update()
            import time; time.sleep(1)
        if not os.path.exists(src_file):
            messagebox.showerror("Lỗi", f"Không tìm thấy file {id_val}.can trong thư mục Download. Hãy chắc chắn đã tải file xong!")
            
            return None
        try:
            hf.copy_file_by_name(download_dir, feature_folder, id_val + ".can")
            self.set_status(f"Đã copy file về: {save_path}", success=True)
            if auto_open:
                os.startfile(save_path)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể copy/mở file: {e}")
            return None
        return save_path

    def copy_downloaded_file(self):
        if not hasattr(self, "_pending_copy_info"):
            messagebox.showwarning("Thiếu thông tin", "Hãy mở script trước để lấy thông tin copy!")
            return
        id_val, feature_folder = self._pending_copy_info
        download_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        src_file = os.path.join(download_dir, id_val + ".can")
        dst_file = os.path.join(feature_folder, id_val + ".can")
        if not os.path.exists(src_file):
            messagebox.showerror("Lỗi", f"Không tìm thấy file {id_val}.can trong thư mục Download. Hãy chắc chắn đã tải file xong!")
            return
        try:
            hf.copy_file_by_name(download_dir, feature_folder, id_val + ".can")
            self.last_save_path = dst_file  # Gán lại ở đây
            self.set_status(f"Đã copy file về: {dst_file}", success=True)
            hf.open_file(dst_file)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể copy file: {e}")

    def set_status(self, message, success=True):
        self.status_label.config(
            text=message,
            fg="#2e7d32" if success else "#c62828",
            bg="#e0e0e0"
        )
        self.status_label.after(4000, lambda: self.status_label.config(text=""))

    def update_open_script_button(self):
        if hasattr(self, "working_dir") and self.working_dir:
            self.btn_open_script.config(style="Success.TButton")
        else:
            self.btn_open_script.config(style="TButton")

    def compare_content_requirement(self):
        # Thêm cột "Check Consistency" nếu chưa có
        if "Check Consistency" not in self.headers:
            self.headers.append("Check Consistency")
            for row in self.data:
                row.append("")

        selected = self.tree.focus()
        if not selected:
            messagebox.showwarning("Chọn dòng", "Hãy chọn một dòng trong bảng!")
            return
        row_idx = self.tree.index(selected)
        row = self.data[row_idx]

        try:
            req_id_idx = self.headers.index("Requirement ID")
            content_idx = self.headers.index("Content Requirement")
            design_idx = self.headers.index("TS_TestDesciption")
            check_idx = self.headers.index("Check Consistency")
            check_TCidx = self.headers.index("Test cases ID")
        except ValueError:
            messagebox.showerror("Lỗi", "Không tìm thấy cột cần thiết!")
            return

        values = row
        save_path = self.ensure_script_file(values, auto_open=False)
        if not save_path or not os.path.exists(save_path):
            row[check_idx] = "CHECK"
            self.refresh_table()
            return

        req_id = str(row[req_id_idx]).strip()
        content_requirement = str(row[content_idx]).strip()
        test_design = str(row[design_idx]).strip()

        # Đọc nội dung file và trích xuất đoạn cần so sánh
        try:
            with open(save_path, "r", encoding="utf-8") as f:
                file_content = f.read()
            start_marker = f"[{req_id}]"
            start_idx = file_content.find(start_marker)
            if start_idx == -1:
                row[check_idx] = "CHECK"
                self.refresh_table()
                return
            end_idx = file_content.find("------------------------", start_idx)
            if end_idx == -1:
                end_idx = len(file_content)
            file_requirement_content = file_content[start_idx:end_idx].strip()

            # Lấy đoạn giữa @design và */
            testcase_id = str(row[check_TCidx]).strip()
            testcase_marker = "*/"
            design_start = file_content.find("@Test design")
            if design_start != -1:
                design_end = file_content.find(testcase_marker, design_start)
                if design_end == -1:
                    design_end = len(file_content)
                file_design_content = file_content[design_start + len("@Test design"):(design_end-len("*/"))].strip()
            else:
                file_design_content = ""
        except Exception:
            row[check_idx] = "CHECK"
            self.refresh_table()
            return

        # Lưu 2 đoạn ra file tạm để so sánh bằng Beyond Compare
        with tempfile.NamedTemporaryFile("w", delete=False, suffix=".txt", encoding="utf-8") as f1, \
            tempfile.NamedTemporaryFile("w", delete=False, suffix=".txt", encoding="utf-8") as f2:
            f1.write(content_requirement+"\n Test design:"+test_design)
            f1_path = f1.name
            f2.write(file_requirement_content+"\n Test design:"+file_design_content)
            f2_path = f2.name

        # So sánh test design
        # with tempfile.NamedTemporaryFile("w", delete=False, suffix=".txt", encoding="utf-8") as f3, \
        #     tempfile.NamedTemporaryFile("w", delete=False, suffix=".txt", encoding="utf-8") as f4:
        #     f3.write(test_design)
        #     f3_path = f3.name
        #     f4.write(file_design_content)
        #     f4_path = f4.name

        # Mở 2 cửa sổ Beyond Compare: 1 cho content requirement, 1 cho test design
        try:
            subprocess.Popen([r'C:\Program Files\Beyond Compare 4\BCompare.exe', f1_path, f2_path])
            # subprocess.Popen([r'C:\Program Files\Beyond Compare 4\BCompare.exe', f3_path, f4_path])
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở Beyond Compare: {e}")
            row[check_idx] = "CHECK"
            self.refresh_table()
            return

        # Hiện popup để bạn chọn kết quả
        status = self.ask_consistency_status()
        row[check_idx] = status
        self.refresh_table()
        self.set_status("Đã kiểm tra xong dòng đã chọn!", success=True)

    def ask_consistency_status(self):
        win = tk.Toplevel(self)
        win.title("Chọn kết quả kiểm tra")
        win.geometry("300x120")
        tk.Label(win, text="Kết quả kiểm tra file này là gì?", font=("Segoe UI", 11)).pack(pady=10)
        result = {"value": None}
        def set_result(val):
            result["value"] = val
            win.destroy()
        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", width=8, command=lambda: set_result("OK")).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="NOK", width=8, command=lambda: set_result("NOK")).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="CHECK", width=8, command=lambda: set_result("CHECK")).pack(side="left", padx=5)
        win.grab_set()
        self.wait_window(win)
        return result["value"]

    def refresh_all(self):
        self.save_excel_table()
        self.save_workspace()
        self.refresh_table()
        self.set_status("Đã làm mới, lưu Excel và workspace!", success=True)

if __name__ == "__main__":
    app = CustomApp()
    app.mainloop()