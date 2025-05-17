import os
import shutil
import stat
from tkinter import messagebox
import subprocess
from openpyxl import load_workbook
import openpyxl
from tkinter import filedialog, messagebox, simpledialog


def copy_file_by_name(app, input_folder, output_folder, filename):
    from file_ops import refresh_table
    """
    Copy a file with the given filename from input_folder to output_folder.
    After copying, make sure the file is writable.
    Return (success: bool, input_path: str)
    """
    input_path = os.path.join(input_folder, filename)
    output_path = os.path.join(output_folder, filename)
    try:
        if not os.path.exists(input_path):
            return False, input_path
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        shutil.copy2(input_path, output_path)
        os.chmod(output_path, stat.S_IWRITE | stat.S_IREAD)
        print(f"Copied '{input_path}' to '{output_path}' and made it writable")
        refresh_table(app)
        return True, input_path
    except Exception as e:
        print(f"Error: {e}")
        return False, input_path

def open_file(filepath):
    try:
        os.startfile(filepath)  # Chỉ dùng được trên Windows
    except Exception as e:
        messagebox.showerror("Error", f"Cannot open file: {e}")

def open_excel_file(input_folder, file_name):
    input_path = os.path.join(input_folder, file_name)
    try:
        if not os.path.exists(input_path):
            return False
        # Mở file Excel bằng subprocess
        subprocess.Popen(['start', 'excel', input_path], shell=True)
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False, input_path

def write_excel_cell_by_label(filepath, row_label, col_label, value, sheet_name=None):
    """
    Ghi giá trị vào file Excel tại vị trí xác định bởi (row_label, col_label).
    row_label: giá trị ở cột đầu tiên (ví dụ: tên hàng)
    col_label: giá trị ở hàng đầu tiên (ví dụ: tên cột)
    """
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Tìm vị trí cột theo col_label (ở hàng 1)
    col_idx = None
    for cell in ws[1]:
        if cell.value == col_label:
            col_idx = cell.column
            break
    if col_idx is None:
        raise ValueError(f"Không tìm thấy cột '{col_label}'")

    # Tìm vị trí hàng theo row_label (ở cột A)
    row_idx = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == row_label:
            row_idx = row[0].row
            break
    if row_idx is None:
        raise ValueError(f"Không tìm thấy hàng '{row_label}'")

    # Ghi giá trị vào ô xác định
    ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(filepath)
    wb.close()

def read_excel_cell_by_label(filepath, row_label, col_label, sheet_name=None):
    """
    Đọc giá trị trong file Excel tại vị trí xác định bởi (row_label, col_label).
    row_label: giá trị ở cột đầu tiên (ví dụ: tên hàng)
    col_label: giá trị ở hàng đầu tiên (ví dụ: tên cột)
    """
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Tìm vị trí cột theo col_label (ở hàng 1)
    col_idx = None
    for cell in ws[1]:
        if cell.value == col_label:
            col_idx = cell.column
            break
    if col_idx is None:
        wb.close()
        raise ValueError(f"Không tìm thấy cột '{col_label}'")

    # Tìm vị trí hàng theo row_label (ở cột A)
    row_idx = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == row_label:
            row_idx = row[0].row
            break
    if row_idx is None:
        wb.close()
        raise ValueError(f"Không tìm thấy hàng '{row_label}'")
    
    # Đọc giá trị tại ô xác định
    value = ws.cell(row=row_idx, column=col_idx).value
    wb.close()
    return value
    
def load_excel_table(app):
    from file_ops import refresh_table
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    app.excel_path = file_path
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # Đảm bảo header là chuỗi, không None, không rỗng, không trùng lặp
    app.headers = []
    for i, cell in enumerate(ws[1]):
        val = cell.value if cell.value not in [None, ""] else f"Cột {i+1}"
        while val in app.headers:
            val += "_1"
        app.headers.append(str(val))
    app.data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        app.data.append(list(row))
    wb.close()
    refresh_table(app)

def save_excel_table(app):
    from file_ops import set_status
    if not app.excel_path or not app.data:
        messagebox.showwarning("Chưa có dữ liệu", "Hãy mở file Excel trước!")
        return
    wb = openpyxl.load_workbook(app.excel_path)
    ws = wb.active
    # Ghi header vào dòng 1
    for j, header in enumerate(app.headers, start=1):
        ws.cell(row=1, column=j, value=header)
    # Ghi dữ liệu vào các dòng tiếp theo
    for i, row in enumerate(app.data, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(app.excel_path)
    wb.close()
    set_status(app,"Đã lưu", "Dữ liệu đã được lưu vào file Excel!")

def open_excel_file(app):
    if not app.excel_path:
        messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
        return
    open_excel_file("", app.excel_path)

def add_column_to_table(app):
    from file_ops import refresh_table
    if not app.excel_path:
        messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
        return
    # Hỏi tên cột mới
    new_col_name = simpledialog.askstring("Đổi tên cột", "Nhập tên cột mới:", initialvalue=f"Cột mới {len(app.headers) + 1}")
    if not new_col_name:
        new_col_name = f"Cột mới {len(app.headers) + 1}"
    # Đảm bảo không trùng tên
    while new_col_name in app.headers:
        new_col_name += "_1"
    app.headers.append(new_col_name)
    for row in app.data:
        row.append("")
    refresh_table(app)

def add_row_to_table(app):
    from file_ops import refresh_table
    if not app.excel_path:
        messagebox.showwarning("Chưa có file", "Hãy mở file Excel trước!")
        return
    new_row = [""] * len(app.headers)
    app.data.append(new_row)
    refresh_table(app)